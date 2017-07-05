'''
Created on 2016-8-24

@author: Administrator
'''

#coding=utf-8

import datetime
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis


def modify1002excel(filename = 'c:/1.xlsx', assets=None, hs300price=None):    
    wb = load_workbook(filename = filename)
    ws = wb['基金资产明细']
    nrows = len(ws.rows)
    ncolumns = len(ws.columns)
    lastdate = ws.cell(row = nrows, column = 1).value
    now = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    
    print(lastdate, now)
#     t1 = time.strptime(lastdate, "%Y%m%d")
#     t2 = time.strptime(now, "%Y%m%d")
    if lastdate >= now:
        print("last date is today")
        return 1;
    
    days = (now - lastdate).days
    
    closeprice = ws.cell(row = nrows, column = 9).value
    fundassets = ws.cell(row = nrows, column = 8).value 
        

    for rowcn in range(0, days):
        print(rowcn)
        oldrow = nrows+rowcn    
        newrow = oldrow+1
        newdate = lastdate + datetime.timedelta(days = 1+rowcn)
        if assets and (rowcn == days-1) :
            closeprice = hs300price
            fundassets = assets 
        print(closeprice, fundassets) 
        ws.cell(row = newrow, column = 9).value = closeprice
        ws.cell(row = newrow, column = 8).value = fundassets
        ws.cell(row = newrow, column = 7).value = "=G%d+E%d*0.002/365" % (oldrow, oldrow)
        ws.cell(row = newrow, column = 6).value = "=F%d+E%d*0.01/365" % (oldrow, oldrow)
        ws.cell(row = newrow, column = 5).value = "=H%d-F%d-G%d" % (newrow, newrow, newrow)
        ws.cell(row = newrow, column = 4).value = "=$D$2"
        ws.cell(row = newrow, column = 3).value = "=I%d/$I$2" % (newrow)
        ws.cell(row = newrow, column = 2).value = "=E%d/D%d" % (newrow, newrow)
        ws.cell(row = newrow, column = 1).value = newdate
        
        for cx in range(1, ncolumns+1) :
            ws.cell(row = newrow, column = cx).number_format = ws.cell(row = oldrow, column = cx).number_format 
       

    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=365)
    
    # Chart with date axis
    c2 = LineChart()
    c2.title = "钱塘1002号基金净值走势图"
    c2.style = 2
    # c2.y_axis.title = "Size"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'm-d'
    c2.x_axis.majorTimeUnit = "months"
  
    # c2.x_axis.title = "Date"
    
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=365)
    c2.set_categories(dates)    
    
#     ws.add_chart(c2, "B%d" % (nrows+2))
    ws1 = wb['净值走势图']
    ws1.add_chart(c2, "B3")
    
    wb.save(filename) 
    os.system(filename) 
    
    return 0


    
if __name__ == '__main__':
    filename = 'c:/1.xlsx'
    
    modify1002excel(filename=filename)
    print("modify over")   


    exit()


    
    wb = Workbook()
    ws = wb.active
    ws = wb.create_sheet() # insert at the end (default)
    ws = wb.create_sheet(0) # insert at first position
    ws.title = "New Title"
    ws.sheet_properties.tabColor = "1072BA"
    ws = wb["New Title"] 
    print(wb.get_sheet_names())
    for sheet in wb:
        print(sheet.title)
    
    c = ws['A4']
    ws['A4'] = 4
    c = ws.cell('A4')
    d = ws.cell(row = 4, column = 2)
    
    cell_range = ws['A1':'C2']
    tuple(ws.iter_rows('A1:C2'))
    ws.rows
    ws.columns
    c.value = '12%'
    
    ws['A1'].number_format
    ws["A1"] = "=SUM(1, 1)"
    
    
    wb.save('new_big_file.xlsx') 
    # print(ws.rows)
    # for row in ws.rows:
    #    for cell in row:
    #        print(cell.value)
    
    # 遍历多个单元格
    for row in ws.iter_rows('A1:D2'):
        for cell in row:
            print(cell)

    wb = load_workbook("files/concatenate.xlsx")
    ws = wb.active
    
    b1 = ws.cell('B1')
    a6 = ws.cell('A6')
    
    assert b1.value == '=CONCATENATE(A1,A2)'
    assert b1.data_type == 'f'
    
    assert a6.value == '=SUM(A4:A5)'
    assert a6.data_type == 'f'
    
    # test iterator
    
    wb = load_workbook("files/concatenate.xlsx", True)
    ws = wb.active
    
    for row in ws.iter_rows():
        for col in row:
            if col.coordinate == 'B1':
                b1 = col
            elif col.coordinate == 'A6':
                a6 = col
    
    assert b1.internal_value == '=CONCATENATE(A1,A2)'
    assert b1.data_type == 'f'
    
    assert a6.internal_value == '=SUM(A4:A5)'
    assert a6.data_type == 'f'
